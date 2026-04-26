from __future__ import annotations

import argparse
import datetime as dt
import os
import sqlite3
from collections import Counter
from pathlib import Path
import sys
from typing import Any, Iterable

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

try:
    from local_db import DEFAULT_DB_PATH
except ModuleNotFoundError:
    DEFAULT_DB_PATH = os.environ.get("JJC_DB_PATH", str(ROOT / "local.db"))


def normalize_column_name(name: str, index: int) -> str:
    raw = (name or "").strip().lower()
    if not raw:
        return f"column_{index}"
    normalized = []
    for ch in raw:
        if ch.isalnum() or ch == "_":
            normalized.append(ch)
        else:
            normalized.append("_")
    candidate = "".join(normalized).strip("_")
    return candidate or f"column_{index}"


def normalize_table_name(name: str) -> str:
    normalized = []
    for ch in name.strip():
        if ch.isalnum() or ch == "_":
            normalized.append(ch)
        else:
            normalized.append("_")
    return "".join(normalized) or "sheet"


def classify_value(value: Any) -> str:
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    if isinstance(value, (dt.datetime, dt.date)):
        return "datetime"
    text = str(value).strip()
    if not text:
        return "null"

    lowered = text.lower()
    if lowered in {"true", "false", "yes", "no", "y", "n"}:
        return "bool"

    try:
        int(text)
        return "int"
    except ValueError:
        pass

    try:
        float(text)
        return "float"
    except ValueError:
        pass

    return "text"


def infer_sqlite_type(values: Iterable[Any]) -> str:
    counts: Counter[str] = Counter(classify_value(v) for v in values)
    counts.pop("null", None)
    if not counts:
        return "TEXT"
    if counts.get("text"):
        return "TEXT"
    if counts.get("datetime"):
        return "TEXT"
    if counts.get("float"):
        return "REAL"
    if counts.get("int") or counts.get("bool"):
        return "INTEGER"
    return "TEXT"


def coerce_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        lowered = stripped.lower()
        if lowered in {"true", "yes", "y"}:
            return 1
        if lowered in {"false", "no", "n"}:
            return 0
        return stripped
    return value


def read_sheet_rows(workbook_path: str, sheet_name: str) -> tuple[list[str], list[tuple[Any, ...]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = workbook[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        headers_raw = next(rows_iter, None)
        if headers_raw is None:
            return [], []

        headers = [normalize_column_name(str(col) if col is not None else "", idx + 1) for idx, col in enumerate(headers_raw)]

        seen: dict[str, int] = {}
        deduped_headers: list[str] = []
        for col in headers:
            if col in seen:
                seen[col] += 1
                deduped_headers.append(f"{col}_{seen[col]}")
            else:
                seen[col] = 1
                deduped_headers.append(col)

        data_rows: list[tuple[Any, ...]] = []
        col_count = len(deduped_headers)
        for row in rows_iter:
            if row is None:
                continue
            trimmed = tuple(row[:col_count]) if len(row) >= col_count else tuple(list(row) + [None] * (col_count - len(row)))
            if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in trimmed):
                continue
            data_rows.append(tuple(coerce_value(cell) for cell in trimmed))
        return deduped_headers, data_rows
    finally:
        workbook.close()


def recreate_table(conn: sqlite3.Connection, table_name: str, columns: list[str], rows: list[tuple[Any, ...]]) -> None:
    if not columns:
        return

    samples_by_col = list(zip(*rows)) if rows else [[] for _ in columns]
    col_defs = []
    for idx, col in enumerate(columns):
        sqlite_type = infer_sqlite_type(samples_by_col[idx] if rows else [])
        col_defs.append(f'"{col}" {sqlite_type}')

    conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
    create_sql = f'CREATE TABLE "{table_name}" ({", ".join(col_defs)});'
    conn.execute(create_sql)

    if not rows:
        return

    placeholders = ", ".join("?" for _ in columns)
    quoted_cols = ", ".join(f'"{col}"' for col in columns)
    insert_sql = f'INSERT INTO "{table_name}" ({quoted_cols}) VALUES ({placeholders})'
    conn.executemany(insert_sql, rows)


def ingest_excel_to_sqlite(excel_path: str, db_path: str) -> None:
    excel_path_obj = Path(excel_path)
    if not excel_path_obj.exists():
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    folder = os.path.dirname(os.path.abspath(db_path))
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet_names = workbook.sheetnames
    finally:
        workbook.close()

    with sqlite3.connect(db_path) as conn:
        conn.execute("PRAGMA foreign_keys = OFF;")
        for sheet in sheet_names:
            table_name = normalize_table_name(sheet)
            columns, rows = read_sheet_rows(excel_path, sheet)
            recreate_table(conn, table_name, columns, rows)
            print(f"Loaded {sheet} -> {table_name}: {len(rows)} rows")
        conn.commit()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Ingest an Excel workbook into SQLite tables.")
    parser.add_argument(
        "--excel",
        default=str(ROOT / "data" / "NovaCarta_EMEA_Dataset_DUMMY.xlsx"),
        help="Path to source .xlsx file.",
    )
    parser.add_argument(
        "--db",
        default=DEFAULT_DB_PATH,
        help="Output SQLite database path.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    excel_path = str(Path(args.excel).resolve())
    db_path = str(Path(args.db).resolve())
    ingest_excel_to_sqlite(excel_path, db_path)
    print(f"Ingestion complete. SQLite DB: {db_path}")


if __name__ == "__main__":
    main()
